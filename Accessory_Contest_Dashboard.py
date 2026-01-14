import streamlit as st
import pandas as pd
from datetime import date, datetime
import hmac
import hashlib
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import io
import time

LOGOUT_URL = "https://totallywirelessgroup.streamlit.app/"

def navigate(page_name):
    st.query_params["page"] = page_name
    st.rerun()

SHARED_SECRET = "TWG_ACCESSORY_BONUS_2026_SECURE"

def validate_token(token):
    try:
        username, timestamp, signature = token.split("|")

        message = f"{username}|{timestamp}"

        expected_signature = hmac.new(
            SHARED_SECRET.encode(),
            message.encode(),
            hashlib.sha256
        ).hexdigest()

        # Signature check
        if not hmac.compare_digest(signature, expected_signature):
            return None

        # Token expiry (10 minutes)
        if time.time() - int(timestamp) > 600:
            return None

        return username

    except Exception:
        return None

# ---------------- PAGE CONFIG ----------------
st.set_page_config(page_title="Accessory Contest Dashboard", layout="wide")

# ---------------- QUERY PARAMS ----------------
qp = st.query_params
token = st.query_params.get("token")

if not token:
    st.error("Unauthorized access.")
    st.stop()

username = validate_token(token)

if not username:
    st.error("Session expired or invalid access.")
    st.stop()
page = qp.get("page", "Home Page")

# ---------------- DATE LOGIC ----------------
today = datetime.today()
THIS_MONTH_START = date(today.year, today.month, 1)
NEXT_MONTH_START = date(today.year + (today.month // 12), ((today.month % 12) + 1), 1)
THIS_MONTH_END = NEXT_MONTH_START - pd.Timedelta(days=1)
THIS_MONTH_LABEL = today.strftime("%b-%y")

# ---------------- FILE ----------------
DATA_FILE = "Accessory_Contest.csv"

# ---------------- LOAD DATA ----------------
df = pd.read_csv(DATA_FILE)
df["Date"] = pd.to_datetime(df["adddate"], errors="coerce").dt.date

df = df[
    [
        "adddate", "marketid", "custno", "company", "item", "itmdesc",
        "qty", "Accessory", "minprice", "Cost", "discount",
        "Profit", "adduser", "Fullname", "invno", "state", "Date"
    ]
]

# ---------------- FILTER BY USER ----------------
def get_user_df():
    if username == "Admin":
        return df.copy()
    return df[df["adduser"] == username].copy()

# ---------------- SIDEBAR ----------------
with st.sidebar:
    st.markdown("### Logged in as")
    st.markdown(f"**{username}**")
    st.divider()

    if st.button("Home Page"):
        navigate("Home Page")

    if st.button("Summary"):
        navigate("Summary")

    if st.button("Detailed"):
        navigate("Detailed")

    st.divider()
# ====================================================
# ==================== HOME PAGE =====================
# ====================================================
if page == "Home Page":

    st.markdown(
        f"""
        <div style="text-align:left;">
            <h1>TWG | TOTALLY WIRELESS GROUP |</h1>
            <h2>Welcome to Accessory Bonus Dashboard!</h2>
            <h3><strong>Top 5 Employees by Bonus – {THIS_MONTH_LABEL}</strong></h3>
            <h3><strong></strong></h3>
        </div>
        """,
        unsafe_allow_html=True
    )

    home_df = df.copy()

    home_df = home_df[
    (home_df["Date"] >= THIS_MONTH_START) &
    (home_df["Date"] <= THIS_MONTH_END)
]

    if home_df.empty:
        st.info("No data available to calculate bonuses.")
    else:
        summary = (
            home_df
            .groupby(["adduser", "Fullname"], as_index=False)
            .agg(
                Total_Accessory=("Accessory", "sum"),
                Total_Profit=("Profit", "sum")
            )
        )

        def calculate_bonus(row):
            acc = row["Total_Accessory"]
            profit = row["Total_Profit"]

            if acc <= 2999:
                pct = 0.08
            elif acc <= 5999:
                pct = 0.10
            elif acc <= 9999:
                pct = 0.15
            else:
                pct = 0.17

            return round(profit * pct, 2)

        summary["Bonus"] = summary.apply(calculate_bonus, axis=1)

        top5 = summary.sort_values("Bonus", ascending=False).head(5).reset_index(drop=True)

        medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]

        for idx, row in top5.iterrows():
            if idx == 0:
                # 1st place – BIG & BOLD
                st.markdown(
                    f"""
                    <div style="text-align:left; font-size:35px; font-weight:800; margin-bottom:12px;">
                        {medals[idx]} {row['adduser']} — {row['Fullname']} — ${row['Bonus']:,.2f}
                    </div>
                    """,
                    unsafe_allow_html=True
                )
            else:
                # Others – normal size
                st.markdown(
                    f"""
                    <div style="text-align:left; font-size:25px; margin-bottom:8px;">
                        {medals[idx]} {row['adduser']} — {row['Fullname']} — ${row['Bonus']:,.2f}
                    </div>
                    """,
                    unsafe_allow_html=True
                )

# ====================================================
# ==================== SUMMARY =======================
# ====================================================
elif page == "Summary":

    st.title("Summary")

    summary_df = get_user_df()

    if summary_df.empty:
        st.info("No data available for the selected user.")
    else:
        fullnames = summary_df["Fullname"].dropna().unique().tolist()
        selected_name = st.selectbox("Select Fullname", ["All"] + fullnames)

        if selected_name != "All":
            summary_df = summary_df[summary_df["Fullname"] == selected_name]

        ntid = summary_df["adduser"].dropna().unique().tolist()
        selected_ntid = st.selectbox("Select NTID", ["All"] + ntid)

        if selected_ntid != "All":
            summary_df = summary_df[summary_df["adduser"] == selected_ntid]

        companies = summary_df["company"].dropna().unique().tolist()
        selected_company = st.selectbox("Select Store", ["All"] + companies)

        if selected_company != "All":
            summary_df = summary_df[summary_df["company"] == selected_company]

        min_date = summary_df["Date"].min()
        max_date = summary_df["Date"].max()

        if pd.notna(min_date) and pd.notna(max_date):
            default_start = max(min_date, THIS_MONTH_START)
            default_end = min(max_date, THIS_MONTH_END)

            start_date, end_date = st.date_input(
                "Select Date Range",
                value=(default_start, default_end),
                min_value=min_date,
                max_value=max_date
            )

            summary_df = summary_df[
                (summary_df["Date"] >= start_date) &
                (summary_df["Date"] <= end_date)
            ]

        total_qty = summary_df["qty"].sum()
        total_accessory = summary_df["Accessory"].sum()
        total_profit = summary_df["Profit"].sum()

        if total_accessory <= 2999:
            tier, bonus_pct = "Tier 1", 0.08
        elif total_accessory <= 5999:
            tier, bonus_pct = "Tier 2", 0.10
        elif total_accessory <= 9999:
            tier, bonus_pct = "Tier 3", 0.15
        else:
            tier, bonus_pct = "Tier 4", 0.17

        bonus = total_profit * bonus_pct

        c1, c2, c3 = st.columns(3)
        c1.metric("Total Qty", total_qty)
        c2.metric("Total Accessory", f"${total_accessory:,.2f}")
        c3.metric("Total Profit", f"${total_profit:,.2f}")

        c4, c5, c6 = st.columns(3)
        c4.metric("Tier", tier)
        c5.metric("Bonus %", f"{bonus_pct*100:.0f}%")
        c6.metric("Bonus", f"${bonus:,.2f}")

        # ================= DOWNLOAD BUTTON =================
        # Function to prepare summary data based on current filtered_df
        def prepare_export(df):
            export_rows = []

            # Group by user & fullname to calculate totals
            grouped = df.groupby(["adduser", "Fullname", "marketid", "company"], as_index=False).agg(
                Total_Qty=("qty", "sum"),
                Total_Accessory=("Accessory", "sum"),
                Total_Profit=("Profit", "sum")
            )

            for _, row in grouped.iterrows():
                # Determine tier and bonus %
                acc = row["Total_Accessory"]
                profit = row["Total_Profit"]

                if acc <= 2999:
                    tier, bonus_pct = "Tier 1", 0.08
                elif acc <= 5999:
                    tier, bonus_pct = "Tier 2", 0.10
                elif acc <= 9999:
                    tier, bonus_pct = "Tier 3", 0.15
                else:
                    tier, bonus_pct = "Tier 4", 0.17

                bonus = profit * bonus_pct

                export_rows.append({
                    "marketid": row["marketid"],
                    "company": row["company"],
                    "adduser": row["adduser"],
                    "Fullname": row["Fullname"],
                    "Total Qty": row["Total_Qty"],
                    "Total Accessory": row["Total_Accessory"],
                    "Total Profit": row["Total_Profit"],
                    "Tier": tier,
                    "Bonus %": f"{bonus_pct*100:.0f}%",
                    "Bonus": round(bonus, 2)
                })

            return export_rows

        # Prepare filtered export data
        export_data = prepare_export(summary_df)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Write headers
        if export_data:
            headers = list(export_data[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Write data rows
            for row_num, row_data in enumerate(export_data, 2):
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col_num, value=row_data[header])
                    ws.column_dimensions[get_column_letter(col_num)].width = max(15, len(header)+2)

    # Save to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Streamlit download button
    st.download_button(
        label="Download Summary",
        data=output,
        file_name=f"Accessory_Summary_{THIS_MONTH_LABEL}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ====================================================
# ==================== DETAILED ======================
# ====================================================
elif page == "Detailed":

    st.title("Detailed Data")

    filtered_df = get_user_df()

    if filtered_df.empty:
        st.info("No data available for the selected user.")
    else:
        fullnames = filtered_df["Fullname"].dropna().unique().tolist()
        selected_name = st.selectbox("Select Fullname", ["All"] + fullnames)

        if selected_name != "All":
            filtered_df = filtered_df[filtered_df["Fullname"] == selected_name]

        ntid = filtered_df["adduser"].dropna().unique().tolist()
        selected_ntid = st.selectbox("Select NTID", ["All"] + ntid)

        if selected_ntid != "All":
            filtered_df = filtered_df[filtered_df["adduser"] == selected_ntid]

        stores = filtered_df["company"].dropna().unique().tolist()
        selected_store = st.selectbox("Select Store", ["All"] + stores)

        if selected_store != "All":
            filtered_df = filtered_df[filtered_df["company"] == selected_store]

        min_date = filtered_df["Date"].min()
        max_date = filtered_df["Date"].max()

        if pd.notna(min_date) and pd.notna(max_date):
            default_start = max(min_date, THIS_MONTH_START)
            default_end = min(max_date, THIS_MONTH_END)

            start_date, end_date = st.date_input(
                "Select Date Range",
                value=(default_start, default_end),
                min_value=min_date,
                max_value=max_date
            )

            filtered_df = filtered_df[
                (filtered_df["Date"] >= start_date) &
                (filtered_df["Date"] <= end_date)
            ]

        # Keep all columns except "Date" for display
        columns_to_display = [col for col in filtered_df.columns if col != "Date"]
        filtered_df_display = filtered_df.loc[:, filtered_df.columns.intersection(columns_to_display)]

        # Format adddate column for display
        if "adddate" in filtered_df_display.columns:
            filtered_df_display = filtered_df_display.copy()
            filtered_df_display["adddate"] = pd.to_datetime(filtered_df_display["adddate"])
            filtered_df_display["adddate"] = filtered_df_display["adddate"].dt.strftime("%m/%d/%Y %I:%M %p")

        # Display the DataFrame
        st.dataframe(filtered_df_display, use_container_width=True)
